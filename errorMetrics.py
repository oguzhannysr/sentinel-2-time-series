import pandas as pd
import scipy.stats as ss
import numpy as np   
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
excel_path = "C:/Users/Lenovo/Downloads/TimeSeries_per25cloud_2022-10-12_aly.xlsx"
excel_output_path = 'D:\\Projects\\BarleyWheat\\Excel_files\\'
ExcelWorkbook = openpyxl.load_workbook(filename = excel_path ,data_only=True)
crop = 'Wheat'
#%%
threshold_cor_range = range(60,90,1)
threshold_mindist_range = range(200,50,-5)
if crop == 'Barley':
    usecols= 'A:AR'
    data = pd.read_excel(excel_path,
                       sheet_name=crop, 
                       usecols=usecols, 
                       skiprows= (lambda x: x<1 or x >2),
                       header=None)
elif crop == 'Wheat':
    usecols= 'A:BF'    
    data = pd.read_excel(excel_path,
                       sheet_name=crop, 
                       usecols= usecols, 
                       skiprows= (lambda x: x<1 or x >2),
                       header=None)
    

veg_name_for = pd.read_excel(excel_path,
                   sheet_name=crop, 
                   usecols= 'A')
veg_name_for = veg_name_for['Unnamed: 0'].unique().astype(str)
veg_name_for = np.delete(veg_name_for, np.argwhere(veg_name_for == 'nan'))


ws = ExcelWorkbook[crop]

#%%
for k,l in zip(threshold_cor_range,threshold_mindist_range):
    e = 15
    f = 77
    d = 0
    p = 2
    #m = 0

    while e < ws.max_row:
        data_rows = []
    
        for row in ws['C{}'.format(e):'{}{}'.format(usecols[2:],f)]:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)
        
        
        df = pd.DataFrame(data_rows)
        df.columns = data.iloc[0][2:].convert_dtypes(int)
        df_cor = df
        df_min = df
        
        detect_cor = pd.DataFrame(columns=df_min.columns,index=np.round(np.arange(0.7,0.89,0.01),4) ,dtype='float64')
        detect_mdst = pd.DataFrame(columns=df_min.columns,index=np.round(np.arange(0.05,0.02,-0.0015),4) ,dtype='float64')
        
        corelation = pd.DataFrame(columns=df_cor.columns,index=[0],dtype='float64')
        mindist = pd.DataFrame(columns=df_min.columns,index=[0],dtype='float64')
        
        detect_corelation = pd.DataFrame()
        detect_mindist = pd.DataFrame()
        
        df_cor2 = pd.DataFrame(columns=df_min.columns,index=np.arange(0,df_cor.shape[1]) ,dtype='float64')
        df_min2 = pd.DataFrame(columns=df_min.columns,index=np.arange(0,df_min.shape[1]) ,dtype='float64')
        
        a = 0
        
        for x in range(0,corelation.shape[1]):
            for i,j in enumerate(df_cor):
                mean = df_cor.mean(axis=1)
                corr,null = ss.pearsonr(df_cor[j],mean)
                corelation.iloc[0,i] = corr
                detect = corelation.idxmin(axis=1)
                detect = detect[0]
                value = corelation.min(axis=1)
                #stp = corelation.isnull().values.any()
        
        
            for i in corelation.columns:
                df_cor2.loc[a][i] = corelation.loc[0][i]
            a = a + 1
            
            if value[0]<k/100:
                drop_parcels = corelation[detect]
                detect_corelation = detect_corelation.append(drop_parcels)
                corelation = corelation.drop(detect, axis=1)
                df_cor = df_cor.drop(detect, axis=1)
                print(value[0]) 
            
       # for q in detect_corelation.index:
      #      detect_cor.loc[k/100][q] = detect_corelation.loc[q].values[0]
            
        a = 0
        for x in range(0,mindist.shape[1]):
            for i,j in enumerate(df_min):
                mean = df_min.mean(axis=1)
                min_dist = (np.linalg.norm(df_min[j] - mean)/len(mean))/10000
                mindist.iloc[0,i] = min_dist
                detect = mindist.idxmax(axis=1)
                detect = detect[0]
                value = mindist.max(axis=1)
                
            for i in mindist.columns:
                df_min2.loc[a][i] = mindist.loc[0][i]
            a = a + 1
                
            if value[0]>(l/10000):
                drop_parcels = mindist[detect]
                detect_mindist = detect_mindist.append(drop_parcels)
                mindist = mindist.drop(detect, axis=1)
                df_min = df_min.drop(detect, axis=1)
                print(value[0])                 
        
        #for q in detect_mindist.index:
        #    detect_mdst.loc[l/10000][q] = detect_mindist.loc[q].values[0]  
            
        e = e + 82
        f = f + 82   
        
        with pd.ExcelWriter(excel_output_path + "/" + "{}_{}_{}_{}.xlsx".format(veg_name_for[d],k,l,crop)) as writer1:
            df_cor2.to_excel(writer1, sheet_name = 'df_cor', index = False)
            df_min2.to_excel(writer1, sheet_name = 'df_min', index = False)
            corelation.to_excel(writer1, sheet_name = 'corelation', index = False)
            mindist.to_excel(writer1, sheet_name = 'mindist', index = False)
            detect_corelation.to_excel(writer1, sheet_name = 'detect_corelation', index = True)
            detect_mindist.to_excel(writer1, sheet_name = 'detect_mindist', index = True)
        d = d + 1
            
    #detect_cor.to_excel(excel_output_path + "/" + "corelation_{}_{}.xlsx".format(crop,veg_name_for[m]))
    #detect_mdst.to_excel(excel_output_path + "/" + "mindist_{}_{}.xlsx".format(crop,veg_name_for[m]))
   #m = m + 1

            
    
